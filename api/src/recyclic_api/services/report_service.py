"""Service for generating bulk exports of cash sessions and reception tickets (Story B45-P1)."""

import csv
from io import BytesIO, StringIO
from typing import List, Optional, Dict, Any
from datetime import datetime
from pathlib import Path

from sqlalchemy.orm import Session, joinedload, selectinload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from recyclic_api.models.cash_session import CashSession
from recyclic_api.models.ticket_depot import TicketDepot
from recyclic_api.models.user import User
from recyclic_api.models.site import Site
from recyclic_api.models.cash_register import CashRegister
from recyclic_api.schemas.cash_session import CashSessionFilters
from recyclic_api.services.cash_session_service import CashSessionService
from recyclic_api.services.reception_service import ReceptionService
from uuid import UUID


def _format_amount(value: Optional[float]) -> str:
    """Format un montant avec virgule comme séparateur décimal (format français)"""
    if value is None:
        return ''
    return f"{value:.2f}".replace('.', ',')


def _format_date(dt: Optional[datetime]) -> str:
    """Format une date en string (YYYY-MM-DD HH:MM:SS)"""
    if dt is None:
        return ''
    return dt.strftime('%Y-%m-%d %H:%M:%S')


def generate_bulk_cash_sessions_csv(
    db: Session,
    filters: CashSessionFilters,
    max_items: int = 10000
) -> BytesIO:
    """
    Génère un export CSV consolidé de toutes les sessions de caisse filtrées.
    
    Args:
        db: Session de base de données
        filters: Filtres à appliquer
        max_items: Nombre maximum d'éléments à exporter (sécurité)
    
    Returns:
        BytesIO contenant le CSV
    """
    service = CashSessionService(db)
    
    # Récupérer toutes les sessions (sans limite de pagination pour export)
    # Mais limiter à max_items pour sécurité
    
    # Utiliser une copie propre pour ne pas modifier l'objet original par effet de bord
    import copy
    filters_export = copy.deepcopy(filters)
    
    # Utiliser object.__setattr__ pour contourner la validation Pydantic sur limit
    object.__setattr__(filters_export, 'limit', max_items)
    object.__setattr__(filters_export, 'skip', 0)
    
    sessions, total = service.get_sessions_with_filters(filters_export)
    
    if total > max_items:
        raise ValueError(f"Trop de sessions à exporter ({total}). Maximum: {max_items}")
    
    # Charger les relations nécessaires
    session_ids = [s.id for s in sessions]
    sessions_with_relations = db.query(CashSession).filter(
        CashSession.id.in_(session_ids)
    ).options(
        joinedload(CashSession.operator),
        joinedload(CashSession.site),
        joinedload(CashSession.register)
    ).all()
    
    # Créer un mapping pour accès rapide
    sessions_map = {s.id: s for s in sessions_with_relations}
    
    buffer = BytesIO()
    # Utiliser utf-8-sig pour BOM UTF-8 (compatibilité Excel)
    writer = csv.writer(StringIO(), delimiter=';', quoting=csv.QUOTE_MINIMAL)
    
    # En-têtes CSV ordonnés selon la demande
    headers = [
        'Date Ouverture',
        'Date Fermeture',
        'Opérateur',
        'Caisse',
        'Site',
        'Montant Initial (€)',
        'Total Ventes (€)',
        'Nombre Ventes',
        'Nombre Articles',
        'Total Dons (€)',
        'Montant Clôture (€)',
        'Montant Réel (€)',
        'Écart (€)',
        'Commentaire Écart',
        'Statut',
        'ID Session'
    ]
    
    # Écrire les en-têtes
    output = StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    
    # Écrire les données
    for session in sessions:
        s = sessions_map.get(session.id, session)
        operator = s.operator if hasattr(s, 'operator') else None
        site = s.site if hasattr(s, 'site') else None
        register = s.register if hasattr(s, 'register') else None
        
        operator_name = ''
        if operator:
            operator_name = (getattr(operator, 'full_name', None) or 
                           getattr(operator, 'username', None) or 
                           getattr(operator, 'telegram_id', None) or '')
        
        site_name = getattr(site, 'name', '') if site else ''
        register_name = getattr(register, 'name', '') if register else ''
        
        # Conversion sécurisée du statut
        status_str = str(s.status.value) if hasattr(s.status, 'value') else str(s.status) if s.status else ''
        
        row = [
            _format_date(s.opened_at),
            _format_date(s.closed_at),
            operator_name,
            register_name,
            site_name,
            _format_amount(s.initial_amount),
            _format_amount(s.total_sales),
            str(s.number_of_sales or 0),
            str(s.total_items or 0),
            _format_amount(s.total_donations),
            _format_amount(s.closing_amount),
            _format_amount(s.actual_amount),
            _format_amount(s.variance),
            s.variance_comment or '',
            status_str,
            str(s.id)
        ]
        writer.writerow(row)
    
    # Convertir en bytes
    csv_content = output.getvalue()
    buffer.write(csv_content.encode('utf-8-sig'))
    buffer.seek(0)
    
    return buffer


def generate_bulk_cash_sessions_excel(
    db: Session,
    filters: CashSessionFilters,
    max_items: int = 10000
) -> BytesIO:
    """
    Génère un export Excel avec onglets "Résumé" et "Détails" pour toutes les sessions filtrées.
    
    Args:
        db: Session de base de données
        filters: Filtres à appliquer
        max_items: Nombre maximum d'éléments à exporter (sécurité)
    
    Returns:
        BytesIO contenant le fichier Excel
    """
    service = CashSessionService(db)
    
    # Récupérer toutes les sessions
    # Créer une copie des filtres avec limit modifié
    import copy
    filters_export = copy.deepcopy(filters)
    
    # Utiliser object.__setattr__ pour contourner la validation Pydantic sur limit
    object.__setattr__(filters_export, 'limit', max_items)
    object.__setattr__(filters_export, 'skip', 0)
    
    sessions, total = service.get_sessions_with_filters(filters_export)
    
    if total > max_items:
        raise ValueError(f"Trop de sessions à exporter ({total}). Maximum: {max_items}")
    
    # Charger les relations nécessaires
    session_ids = [s.id for s in sessions]
    sessions_with_relations = db.query(CashSession).filter(
        CashSession.id.in_(session_ids)
    ).options(
        joinedload(CashSession.operator),
        joinedload(CashSession.site),
        joinedload(CashSession.register)
    ).all()
    
    sessions_map = {s.id: s for s in sessions_with_relations}
    
    buffer = BytesIO()
    wb = Workbook()
    
    # Supprimer la feuille par défaut si elle existe
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # === ONGLET RÉSUMÉ ===
    ws_summary = wb.create_sheet("Résumé")
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes résumé
    summary_headers = [
        'Date Ouverture', 'Opérateur', 'Caisse', 'Site',
        'CA Total (€)', 'Nb Ventes', 'Nb Articles', 'Écart (€)', 'Statut'
    ]
    ws_summary.append(summary_headers)
    
    # Appliquer style aux en-têtes
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données résumé
    total_ca = 0.0
    total_ventes = 0
    total_articles = 0
    
    for session in sessions:
        s = sessions_map.get(session.id, session)
        operator = s.operator if hasattr(s, 'operator') else None
        site = s.site if hasattr(s, 'site') else None
        register = s.register if hasattr(s, 'register') else None
        
        operator_name = ''
        if operator:
            operator_name = (getattr(operator, 'full_name', None) or 
                           getattr(operator, 'username', None) or 
                           getattr(operator, 'telegram_id', None) or '')
        
        site_name = getattr(site, 'name', '') if site else ''
        register_name = getattr(register, 'name', '') if register else ''
        
        # Conversion sécurisée du statut
        status_str = str(s.status.value) if hasattr(s.status, 'value') else str(s.status) if s.status else ''
        
        ca = s.total_sales or 0.0
        nb_ventes = s.number_of_sales or 0
        nb_articles = s.total_items or 0
        
        total_ca += ca
        total_ventes += nb_ventes
        total_articles += nb_articles
        
        row = [
            _format_date(s.opened_at),
            operator_name,
            register_name,
            site_name,
            _format_amount(ca),
            str(nb_ventes),
            str(nb_articles),
            _format_amount(s.variance),
            status_str
        ]
        ws_summary.append(row)
    
    # Ligne de totaux
    ws_summary.append([
        'TOTAL',
        '',
        '',
        '',
        _format_amount(total_ca),
        str(total_ventes),
        str(total_articles),
        '',
        ''
    ])
    
    # Style ligne totaux
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    
    # Largeurs colonnes résumé
    ws_summary.column_dimensions['A'].width = 20 # Date
    ws_summary.column_dimensions['B'].width = 25 # Opérateur
    ws_summary.column_dimensions['C'].width = 20 # Caisse
    ws_summary.column_dimensions['D'].width = 20 # Site
    ws_summary.column_dimensions['E'].width = 15 # CA
    ws_summary.column_dimensions['F'].width = 12 # Nb Ventes
    ws_summary.column_dimensions['G'].width = 12 # Nb Articles
    ws_summary.column_dimensions['H'].width = 15 # Ecart
    ws_summary.column_dimensions['I'].width = 12 # Statut
    
    # === ONGLET DÉTAILS ===
    ws_details = wb.create_sheet("Détails")
    
    # En-têtes détails ordonnés
    detail_headers = [
        'Date Ouverture', 'Date Fermeture', 'Opérateur', 'Caisse', 'Site',
        'Montant Initial (€)', 'Total Ventes (€)', 'Nb Ventes', 'Nb Articles',
        'Total Dons (€)', 'Montant Clôture (€)', 'Montant Réel (€)',
        'Écart (€)', 'Commentaire Écart', 'Statut', 'ID Session'
    ]
    ws_details.append(detail_headers)
    
    # Appliquer style aux en-têtes
    for cell in ws_details[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données détails
    for session in sessions:
        s = sessions_map.get(session.id, session)
        operator = s.operator if hasattr(s, 'operator') else None
        site = s.site if hasattr(s, 'site') else None
        register = s.register if hasattr(s, 'register') else None
        
        operator_name = ''
        if operator:
            operator_name = (getattr(operator, 'full_name', None) or 
                           getattr(operator, 'username', None) or 
                           getattr(operator, 'telegram_id', None) or '')
        
        site_name = getattr(site, 'name', '') if site else ''
        register_name = getattr(register, 'name', '') if register else ''
        
        # Conversion sécurisée du statut
        status_str = str(s.status.value) if hasattr(s.status, 'value') else str(s.status) if s.status else ''
        
        row = [
            _format_date(s.opened_at),
            _format_date(s.closed_at),
            operator_name,
            register_name,
            site_name,
            _format_amount(s.initial_amount),
            _format_amount(s.total_sales),
            str(s.number_of_sales or 0),
            str(s.total_items or 0),
            _format_amount(s.total_donations),
            _format_amount(s.closing_amount),
            _format_amount(s.actual_amount),
            _format_amount(s.variance),
            s.variance_comment or '',
            status_str,
            str(s.id)
        ]
        ws_details.append(row)
    
    # Largeurs colonnes détails
    ws_details.column_dimensions['A'].width = 20 # Date Ouv
    ws_details.column_dimensions['B'].width = 20 # Date Ferm
    ws_details.column_dimensions['C'].width = 25 # Operateur
    ws_details.column_dimensions['D'].width = 20 # Caisse
    ws_details.column_dimensions['E'].width = 20 # Site
    ws_details.column_dimensions['F'].width = 15 # Mnt Init
    ws_details.column_dimensions['G'].width = 15 # Tot Ventes
    ws_details.column_dimensions['H'].width = 12 # Nb Ventes
    ws_details.column_dimensions['I'].width = 12 # Nb Articles
    ws_details.column_dimensions['J'].width = 15 # Tot Dons
    ws_details.column_dimensions['K'].width = 15 # Mnt Clot
    ws_details.column_dimensions['L'].width = 15 # Mnt Reel
    ws_details.column_dimensions['M'].width = 15 # Ecart
    ws_details.column_dimensions['N'].width = 30 # Comm
    ws_details.column_dimensions['O'].width = 12 # Statut
    ws_details.column_dimensions['P'].width = 38 # UUID
    
    # Sauvegarder dans le buffer
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def _format_weight(value: Optional[float]) -> str:
    """Format un poids avec virgule comme séparateur décimal (format français)"""
    if value is None:
        return ''
    return f"{value:.3f}".replace('.', ',')


def generate_bulk_reception_tickets_csv(
    db: Session,
    status: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    benevole_id: Optional[UUID] = None,
    search: Optional[str] = None,
    include_empty: bool = False,
    max_items: int = 10000
) -> BytesIO:
    """
    Génère un export CSV détaillé de tous les tickets de réception filtrés.
    Une ligne par ligne de dépôt (LigneDepot).
    
    Args:
        db: Session de base de données
        status: Statut du ticket (opened/closed)
        date_from: Date de début
        date_to: Date de fin
        benevole_id: ID du bénévole
        search: Recherche textuelle
        include_empty: Inclure les tickets vides
        max_items: Nombre maximum d'éléments à exporter (sécurité)
    
    Returns:
        BytesIO contenant le CSV
    """
    from sqlalchemy.orm import selectinload
    from recyclic_api.models.ticket_depot import TicketDepot
    from recyclic_api.models.ligne_depot import LigneDepot
    
    service = ReceptionService(db)
    
    # Récupérer tous les tickets avec leurs lignes et catégories chargées
    tickets, total = service.get_tickets_list(
        page=1,
        per_page=max_items,
        status=status,
        date_from=date_from,
        date_to=date_to,
        benevole_id=benevole_id,
        search=search,
        include_empty=include_empty
    )
    
    if total > max_items:
        raise ValueError(f"Trop de tickets à exporter ({total}). Maximum: {max_items}")
    
    # Recharger les tickets avec les lignes et catégories pour éviter N+1
    ticket_ids = [ticket.id for ticket in tickets]
    tickets_with_details = db.query(TicketDepot).options(
        selectinload(TicketDepot.benevole),
        selectinload(TicketDepot.lignes).selectinload(LigneDepot.category)
    ).filter(TicketDepot.id.in_(ticket_ids)).all()
    
    buffer = BytesIO()
    output = StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    
    # En-têtes CSV - Détail (une ligne par ligne de dépôt)
    headers = [
        'ticket_id',
        'poste_id',
        'ticket_status',
        'ticket_created_at',
        'ticket_closed_at',
        'benevole_username',
        'ticket_total_poids_kg',
        'ticket_total_lignes',
        'ligne_id',
        'category_id',
        'category_label',
        'destination',
        'poids_kg',
        'notes'
    ]
    
    writer.writerow(headers)
    
    # Écrire les données - une ligne par ligne de dépôt
    for ticket in tickets_with_details:
        total_lignes, total_poids = service._calculate_ticket_totals(ticket)
        
        benevole_name = ''
        if ticket.benevole:
            benevole_name = (getattr(ticket.benevole, 'full_name', None) or 
                           getattr(ticket.benevole, 'username', None) or 
                           getattr(ticket.benevole, 'telegram_id', None) or '')
        
        # Si le ticket n'a pas de lignes, créer quand même une ligne pour le ticket
        if not ticket.lignes:
            row = [
                str(ticket.id),
                str(ticket.poste_id),
                ticket.status,
                _format_date(ticket.created_at),
                _format_date(ticket.closed_at),
                benevole_name,
                _format_weight(float(total_poids)),
                str(total_lignes),
                '',  # ligne_id
                '',  # category_id
                '',  # category_label
                '',  # destination
                '',  # poids_kg
                ''   # notes
            ]
            writer.writerow(row)
        else:
            # Une ligne par ligne de dépôt
            for ligne in ticket.lignes:
                category_label = ''
                category_id = ''
                if ligne.category:
                    category_label = ligne.category.name or ''
                    category_id = str(ligne.category.id) if ligne.category.id else ''
                
                destination_value = ''
                if ligne.destination:
                    destination_value = ligne.destination.value if hasattr(ligne.destination, 'value') else str(ligne.destination)
                
                notes = (ligne.notes or '').replace('\n', ' ').replace('\r', ' ').strip()
                
                row = [
                    str(ticket.id),
                    str(ticket.poste_id),
                    ticket.status,
                    _format_date(ticket.created_at),
                    _format_date(ticket.closed_at),
                    benevole_name,
                    _format_weight(float(total_poids)),
                    str(total_lignes),
                    str(ligne.id),
                    category_id,
                    category_label,
                    destination_value,
                    _format_weight(float(ligne.poids_kg)) if ligne.poids_kg else '',
                    notes
                ]
                writer.writerow(row)
    
    # Convertir en bytes
    csv_content = output.getvalue()
    buffer.write(csv_content.encode('utf-8-sig'))
    buffer.seek(0)
    
    return buffer


def generate_bulk_reception_tickets_excel(
    db: Session,
    status: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    benevole_id: Optional[UUID] = None,
    search: Optional[str] = None,
    include_empty: bool = False,
    max_items: int = 10000
) -> BytesIO:
    """
    Génère un export Excel avec onglets "Résumé" et "Détails" pour tous les tickets filtrés.
    
    Args:
        db: Session de base de données
        status: Statut du ticket (opened/closed)
        date_from: Date de début
        date_to: Date de fin
        benevole_id: ID du bénévole
        search: Recherche textuelle
        include_empty: Inclure les tickets vides
        max_items: Nombre maximum d'éléments à exporter (sécurité)
    
    Returns:
        BytesIO contenant le fichier Excel
    """
    service = ReceptionService(db)
    
    # Récupérer tous les tickets
    tickets, total = service.get_tickets_list(
        page=1,
        per_page=max_items,
        status=status,
        date_from=date_from,
        date_to=date_to,
        benevole_id=benevole_id,
        search=search,
        include_empty=include_empty
    )
    
    if total > max_items:
        raise ValueError(f"Trop de tickets à exporter ({total}). Maximum: {max_items}")
    
    buffer = BytesIO()
    wb = Workbook()
    
    # Supprimer la feuille par défaut si elle existe
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # === ONGLET RÉSUMÉ ===
    ws_summary = wb.create_sheet("Résumé")
    
    # Style des en-têtes
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # En-têtes résumé (inclure ID Ticket et Poste ID pour correspondre au CSV)
    summary_headers = [
        'ID Ticket', 'Statut', 'Date Création', 'Date Fermeture',
        'Bénévole', 'Poste ID', 'Nb Lignes', 'Poids Total (kg)'
    ]
    ws_summary.append(summary_headers)
    
    # Appliquer style aux en-têtes
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données résumé
    total_lignes = 0
    total_poids = 0.0
    
    for ticket in tickets:
        nb_lignes, poids = service._calculate_ticket_totals(ticket)
        
        benevole_name = ''
        if ticket.benevole:
            benevole_name = (getattr(ticket.benevole, 'full_name', None) or 
                           getattr(ticket.benevole, 'username', None) or 
                           getattr(ticket.benevole, 'telegram_id', None) or '')
        
        total_lignes += nb_lignes
        total_poids += float(poids)
        
        row = [
            str(ticket.id),
            ticket.status,
            _format_date(ticket.created_at),
            _format_date(ticket.closed_at),
            benevole_name,
            str(ticket.poste_id),
            str(nb_lignes),
            _format_weight(float(poids))
        ]
        ws_summary.append(row)
    
    # Ligne de totaux
    ws_summary.append([
        'TOTAL',
        '',
        '',
        '',
        '',
        '',
        str(total_lignes),
        _format_weight(total_poids)
    ])
    
    # Style ligne totaux
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    
    # Largeurs colonnes résumé
    ws_summary.column_dimensions['A'].width = 38  # UUID
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 20
    ws_summary.column_dimensions['D'].width = 20
    ws_summary.column_dimensions['E'].width = 25
    ws_summary.column_dimensions['F'].width = 38  # Poste ID UUID
    ws_summary.column_dimensions['G'].width = 12
    ws_summary.column_dimensions['H'].width = 18
    
    # === ONGLET DÉTAILS ===
    ws_details = wb.create_sheet("Détail")
    
    # Recharger les tickets avec les lignes et catégories pour éviter N+1
    from recyclic_api.models.ticket_depot import TicketDepot
    from recyclic_api.models.ligne_depot import LigneDepot
    
    ticket_ids = [ticket.id for ticket in tickets]
    tickets_with_details = db.query(TicketDepot).options(
        selectinload(TicketDepot.benevole),
        selectinload(TicketDepot.lignes).selectinload(LigneDepot.category)
    ).filter(TicketDepot.id.in_(ticket_ids)).all()
    
    # En-têtes détails - une ligne par ligne de dépôt
    detail_headers = [
        'ticket_id',
        'poste_id',
        'ticket_status',
        'ticket_created_at',
        'ticket_closed_at',
        'benevole_username',
        'ticket_total_poids_kg',
        'ticket_total_lignes',
        'ligne_id',
        'category_id',
        'category_label',
        'destination',
        'poids_kg',
        'notes'
    ]
    ws_details.append(detail_headers)
    
    # Appliquer style aux en-têtes
    for cell in ws_details[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Données détails - une ligne par ligne de dépôt
    for ticket in tickets_with_details:
        total_lignes, total_poids = service._calculate_ticket_totals(ticket)
        
        benevole_name = ''
        if ticket.benevole:
            benevole_name = (getattr(ticket.benevole, 'full_name', None) or 
                           getattr(ticket.benevole, 'username', None) or 
                           getattr(ticket.benevole, 'telegram_id', None) or '')
        
        # Si le ticket n'a pas de lignes, créer quand même une ligne pour le ticket
        if not ticket.lignes:
            row = [
                str(ticket.id),
                str(ticket.poste_id),
                ticket.status,
                _format_date(ticket.created_at),
                _format_date(ticket.closed_at),
                benevole_name,
                _format_weight(float(total_poids)),
                str(total_lignes),
                '',  # ligne_id
                '',  # category_id
                '',  # category_label
                '',  # destination
                '',  # poids_kg
                ''   # notes
            ]
            ws_details.append(row)
        else:
            # Une ligne par ligne de dépôt
            for ligne in ticket.lignes:
                category_label = ''
                category_id = ''
                if ligne.category:
                    category_label = ligne.category.name or ''
                    category_id = str(ligne.category.id) if ligne.category.id else ''
                
                destination_value = ''
                if ligne.destination:
                    destination_value = ligne.destination.value if hasattr(ligne.destination, 'value') else str(ligne.destination)
                
                notes = (ligne.notes or '').replace('\n', ' ').replace('\r', ' ').strip()
                
                row = [
                    str(ticket.id),
                    str(ticket.poste_id),
                    ticket.status,
                    _format_date(ticket.created_at),
                    _format_date(ticket.closed_at),
                    benevole_name,
                    _format_weight(float(total_poids)),
                    str(total_lignes),
                    str(ligne.id),
                    category_id,
                    category_label,
                    destination_value,
                    _format_weight(float(ligne.poids_kg)) if ligne.poids_kg else '',
                    notes
                ]
                ws_details.append(row)
    
    # Largeurs colonnes détails
    ws_details.column_dimensions['A'].width = 38  # ticket_id UUID
    ws_details.column_dimensions['B'].width = 38  # poste_id UUID
    ws_details.column_dimensions['C'].width = 12  # ticket_status
    ws_details.column_dimensions['D'].width = 20  # ticket_created_at
    ws_details.column_dimensions['E'].width = 20  # ticket_closed_at
    ws_details.column_dimensions['F'].width = 25  # benevole_username
    ws_details.column_dimensions['G'].width = 18  # ticket_total_poids_kg
    ws_details.column_dimensions['H'].width = 12  # ticket_total_lignes
    ws_details.column_dimensions['I'].width = 38  # ligne_id UUID
    ws_details.column_dimensions['J'].width = 38  # category_id UUID
    ws_details.column_dimensions['K'].width = 30  # category_label
    ws_details.column_dimensions['L'].width = 15  # destination
    ws_details.column_dimensions['M'].width = 12  # poids_kg
    ws_details.column_dimensions['N'].width = 40  # notes
    
    # Sauvegarder dans le buffer
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
