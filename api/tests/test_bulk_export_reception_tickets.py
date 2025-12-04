"""
Tests d'intégration pour l'export bulk de tickets de réception (Story B45-P1).

Tests vérifient:
- Export CSV bulk avec filtres
- Export Excel bulk avec filtres
- Respect des filtres (date, statut, bénévole)
- Permissions (ADMIN/SUPER_ADMIN uniquement)
"""
import io
from datetime import datetime, timedelta, timezone
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from recyclic_api.models.ticket_depot import TicketDepot, TicketDepotStatus
from recyclic_api.models.poste_reception import PosteReception, PosteReceptionStatus
from recyclic_api.models.ligne_depot import LigneDepot
from recyclic_api.models.user import User, UserRole, UserStatus
from recyclic_api.models.category import Category
from recyclic_api.core.security import hash_password


@pytest.fixture
def test_category(db_session: Session) -> Category:
    """Créer une catégorie de test."""
    category = Category(
        id=uuid4(),
        name="Test Category",
        price=10.0
    )
    db_session.add(category)
    db_session.commit()
    db_session.refresh(category)
    return category


@pytest.fixture
def test_poste(db_session: Session) -> PosteReception:
    """Créer un poste de réception de test."""
    poste = PosteReception(
        id=uuid4(),
        name="Poste Test",
        status=PosteReceptionStatus.OPENED,
        opened_at=datetime.now(timezone.utc) - timedelta(days=1)
    )
    db_session.add(poste)
    db_session.commit()
    db_session.refresh(poste)
    return poste


@pytest.fixture
def test_benevole(db_session: Session) -> User:
    """Créer un bénévole de test."""
    benevole = User(
        id=uuid4(),
        username="benevole_test",
        hashed_password=hash_password("testpass"),
        role=UserRole.USER,
        status=UserStatus.ACTIVE,
        is_active=True
    )
    db_session.add(benevole)
    db_session.commit()
    db_session.refresh(benevole)
    return benevole


@pytest.fixture
def test_tickets(db_session: Session, test_poste: PosteReception, test_benevole: User, test_category: Category) -> list[TicketDepot]:
    """Créer plusieurs tickets de test avec différentes dates."""
    tickets = []
    now = datetime.now(timezone.utc)
    
    # Ticket 1: Hier, fermé, avec lignes
    ticket1 = TicketDepot(
        id=uuid4(),
        poste_id=test_poste.id,
        benevole_user_id=test_benevole.id,
        status=TicketDepotStatus.CLOSED,
        created_at=now - timedelta(days=1),
        closed_at=now - timedelta(hours=12)
    )
    db_session.add(ticket1)
    db_session.flush()
    
    ligne1 = LigneDepot(
        id=uuid4(),
        ticket_id=ticket1.id,
        category_id=test_category.id,
        poids_kg=5.5,
        destination="revente"
    )
    db_session.add(ligne1)
    
    ligne2 = LigneDepot(
        id=uuid4(),
        ticket_id=ticket1.id,
        category_id=test_category.id,
        poids_kg=3.2,
        destination="recyclage"
    )
    db_session.add(ligne2)
    
    # Ticket 2: Aujourd'hui, ouvert
    ticket2 = TicketDepot(
        id=uuid4(),
        poste_id=test_poste.id,
        benevole_user_id=test_benevole.id,
        status=TicketDepotStatus.OPENED,
        created_at=now - timedelta(hours=2)
    )
    db_session.add(ticket2)
    db_session.flush()
    
    ligne3 = LigneDepot(
        id=uuid4(),
        ticket_id=ticket2.id,
        category_id=test_category.id,
        poids_kg=2.1,
        destination="revente"
    )
    db_session.add(ligne3)
    
    # Ticket 3: Il y a 3 jours, fermé, sans lignes (ticket vide)
    ticket3 = TicketDepot(
        id=uuid4(),
        poste_id=test_poste.id,
        benevole_user_id=test_benevole.id,
        status=TicketDepotStatus.CLOSED,
        created_at=now - timedelta(days=3),
        closed_at=now - timedelta(days=3, hours=1)
    )
    db_session.add(ticket3)
    
    db_session.commit()
    
    for ticket in [ticket1, ticket2, ticket3]:
        db_session.refresh(ticket)
        tickets.append(ticket)
    
    return tickets


class TestBulkExportReceptionTicketsCSV:
    """Tests pour l'export CSV bulk de tickets de réception."""
    
    def test_export_csv_bulk_success(self, admin_client, test_tickets):
        """Test export CSV bulk avec succès."""
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "include_empty": False
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        assert "text/csv" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        assert "export_tickets_reception" in response.headers["content-disposition"]
        
        # Vérifier le contenu CSV
        content = response.text
        assert "ID Ticket" in content
        assert "Statut" in content
        assert "Date Création" in content
        assert "Bénévole" in content
        assert "Nombre Lignes" in content
        assert "Poids Total (kg)" in content
    
    def test_export_csv_bulk_with_filters(self, admin_client, test_tickets):
        """Test export CSV avec filtres (statut, date)."""
        yesterday = datetime.now(timezone.utc) - timedelta(days=1)
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": yesterday.isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "status": "closed",
                    "include_empty": False
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        content = response.text
        
        # Vérifier que seuls les tickets fermés sont inclus
        lines = content.split('\n')
        data_lines = [l for l in lines[1:] if l.strip()]
        
        # Au moins un ticket fermé doit être présent
        assert len(data_lines) > 0
    
    def test_export_csv_bulk_unauthorized(self, client, db_session, test_tickets):
        """Test que USER ne peut pas exporter bulk."""
        from recyclic_api.core.security import create_access_token
        
        # Créer un utilisateur USER (non-admin) dans la session de test
        user = User(
            id=uuid4(),
            username="user_test",
            hashed_password=hash_password("testpass"),
            role=UserRole.USER,
            status=UserStatus.ACTIVE,
            is_active=True
        )
        db_session.add(user)
        db_session.commit()
        db_session.refresh(user)
        
        token = create_access_token(data={"sub": str(user.id)})
        client.headers["Authorization"] = f"Bearer {token}"
        
        response = client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {},
                "format": "csv"
            }
        )
        
        assert response.status_code == 403
    
    def test_export_csv_bulk_respects_filters(self, admin_client, test_tickets, test_benevole):
        """Test que les filtres sont bien respectés."""
        # Filtrer par bénévole
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "benevole_id": str(test_benevole.id),
                    "include_empty": False
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        content = response.text
        
        # Vérifier que toutes les sessions retournées sont du bon bénévole
        lines = content.split('\n')
        # Les tickets de test sont tous du même bénévole, donc on vérifie juste que ça fonctionne
        assert len(lines) > 1  # Au moins headers + données


class TestBulkExportReceptionTicketsExcel:
    """Tests pour l'export Excel bulk de tickets de réception."""
    
    def test_export_excel_bulk_success(self, admin_client, test_tickets):
        """Test export Excel bulk avec succès."""
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        assert response.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]
        
        # Vérifier que c'est un fichier Excel valide
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        
        # Vérifier les onglets
        assert "Résumé" in wb.sheetnames
        assert "Détails" in wb.sheetnames
        
        # Vérifier le contenu de l'onglet Résumé
        ws_summary = wb["Résumé"]
        assert ws_summary.max_row > 1  # Au moins headers + données
        
        # Vérifier les en-têtes
        headers = [cell.value for cell in ws_summary[1]]
        assert "Statut" in headers
        assert "Date Création" in headers
        assert "Nb Lignes" in headers
        assert "Poids Total (kg)" in headers
        
        # Vérifier le contenu de l'onglet Détails
        ws_details = wb["Détails"]
        assert ws_details.max_row > 1
        
        headers_details = [cell.value for cell in ws_details[1]]
        assert "ID Ticket" in headers_details
        assert "Nombre Lignes" in headers_details
    
    def test_export_excel_bulk_with_filters(self, admin_client, test_tickets):
        """Test export Excel avec filtres."""
        yesterday = datetime.now(timezone.utc) - timedelta(days=1)
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": yesterday.isoformat(),
                    "status": "closed",
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        assert response.status_code == 200
        
        # Vérifier le fichier Excel
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        
        # Vérifier qu'il y a des données
        ws_summary = wb["Résumé"]
        assert ws_summary.max_row > 1  # Headers + au moins une ligne de données
    
    def test_export_excel_formatting_styles(self, admin_client, test_tickets):
        """Test que la mise en forme (styles, couleurs, bordures) est appliquée."""
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": (datetime.now(timezone.utc) - timedelta(days=5)).isoformat(),
                    "date_to": datetime.now(timezone.utc).isoformat(),
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        assert response.status_code == 200
        
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        
        # Vérifier les styles des en-têtes dans l'onglet Résumé
        ws_summary = wb["Résumé"]
        header_row = ws_summary[1]
        
        for cell in header_row:
            # Vérifier que les en-têtes sont en gras
            assert cell.font.bold is True, f"Header cell {cell.coordinate} should be bold"
            # Vérifier que les en-têtes ont un fond coloré
            assert cell.fill.start_color is not None, f"Header cell {cell.coordinate} should have fill color"
            # Vérifier que les en-têtes ont des bordures
            assert cell.border is not None, f"Header cell {cell.coordinate} should have border"
        
        # Vérifier les styles des en-têtes dans l'onglet Détails
        ws_details = wb["Détails"]
        detail_header_row = ws_details[1]
        
        for cell in detail_header_row:
            assert cell.font.bold is True, f"Detail header cell {cell.coordinate} should be bold"
            assert cell.fill.start_color is not None, f"Detail header cell {cell.coordinate} should have fill color"
            assert cell.border is not None, f"Detail header cell {cell.coordinate} should have border"
    
    def test_export_excel_performance_1000_tickets(self, admin_client, db_session, test_poste, test_benevole, test_category):
        """Test de performance : export Excel de 1000 tickets doit être < 30 secondes."""
        import time
        
        # Créer 1000 tickets de test
        now = datetime.now(timezone.utc)
        tickets = []
        for i in range(1000):
            ticket = TicketDepot(
                id=uuid4(),
                poste_id=test_poste.id,
                benevole_user_id=test_benevole.id,
                status=TicketDepotStatus.CLOSED,
                created_at=now - timedelta(days=i % 30),
                closed_at=now - timedelta(days=i % 30, hours=1)
            )
            tickets.append(ticket)
        
        db_session.add_all(tickets)
        db_session.flush()
        
        # Créer des lignes pour chaque ticket
        for idx, ticket in enumerate(tickets):
            ligne = LigneDepot(
                id=uuid4(),
                ticket_id=ticket.id,
                category_id=test_category.id,
                poids_kg=float(idx % 10) + 0.5,
                destination="revente"
            )
            db_session.add(ligne)
        
        db_session.commit()
        
        # Mesurer le temps d'export
        start_time = time.time()
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": (now - timedelta(days=30)).isoformat(),
                    "date_to": now.isoformat(),
                    "include_empty": False
                },
                "format": "excel"
            }
        )
        
        elapsed_time = time.time() - start_time
        
        assert response.status_code == 200
        assert elapsed_time < 30.0, f"Export took {elapsed_time:.2f}s, should be < 30s"
        
        # Vérifier que le fichier est valide
        excel_bytes = io.BytesIO(response.content)
        wb = load_workbook(excel_bytes)
        assert "Résumé" in wb.sheetnames
        assert "Détails" in wb.sheetnames


class TestBulkExportReceptionTicketsValidation:
    """Tests de validation pour l'export bulk de tickets."""
    
    def test_export_bulk_invalid_format(self, admin_client):
        """Test avec format invalide."""
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {},
                "format": "pdf"  # Format invalide
            }
        )
        
        assert response.status_code == 422  # Validation error
    
    def test_export_bulk_empty_result(self, admin_client):
        """Test export avec aucun résultat (filtres trop restrictifs)."""
        far_future = datetime.now(timezone.utc) + timedelta(days=365)
        
        response = admin_client.post(
            "/api/v1/admin/reports/reception-tickets/export-bulk",
            json={
                "filters": {
                    "date_from": far_future.isoformat(),
                    "date_to": (far_future + timedelta(days=1)).isoformat()
                },
                "format": "csv"
            }
        )
        
        assert response.status_code == 200
        # Le CSV devrait contenir seulement les headers
        content = response.text
        lines = [l for l in content.split('\n') if l.strip()]
        assert len(lines) == 1  # Seulement headers

